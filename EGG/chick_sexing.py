import pandas as pd
import datetime
from pyx.xl import *
import re
from openpyxl import Workbook, load_workbook
from EGG.hatchery import *

from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl import drawing

#from openpyxl.worksheet.page import PageSetup
def fit(ws, dim=0, orientation='landscape'):#, paperSize=PageSetup.PAPERSIZE_A4):
	ws.page_setup.orientation = orientation
	#ws.page_setup.paperSize = PageSetup.PAPERSIZE_A4
	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.page_setup.fitToWidth = (dim+1) % 2
	ws.page_setup.fitToHeight = dim

def gerar_programa_sexagem(sh, db, date):
	#print(sh['N10942'])


	wb = Workbook()
	ws = wb['Sheet']

	lotes = wb.create_sheet(title='LOTES')


	row = 5

	merge_range(ws, row, 1, row, 10, 'PROGRAMAÇÃO DE SEXAGEM - INCUBATÓRIO RIO CLARO')

	merge_range(ws, row + 3, 8, row + 3, 10, 'Carregamento')
	for i, x in enumerate(['Dia da semana', 'Nascimento', 'Fêmea', 'Macho', 'Pintos mistos previsto', 'Sexadores', 'Início da sexagem']):
		merge_range(ws, row + 3, i + 1, row + 4, i + 1, x)

	set_row(ws, row + 4, ['Data de saída', 'Horário de saída', 'Cliente'], 8)

	row += 5

	main_date = date #datetime.datetime(2024, 3, 4)
	srcrow = 16982	#502
	dif = 56 #45
	_21dayslater = (main_date + datetime.timedelta(days=21)).isocalendar()
	while True:
		curdat = pd.to_datetime(sh[f'N{srcrow}'].value).isocalendar()
		if _21dayslater.year == curdat.year and _21dayslater.week == curdat.week:
			break
		srcrow += dif

	srcrow -= dif * 2

	srcrow -= 2

	for i in range(7):
		srcrow += dif
		hd = pd.to_datetime(sh[f'N{srcrow + 2}'].value)
		print(hd)
		ws[f'A{row}'].value = ['Domingo', 'Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira', 'Sábado'][hd.isocalendar().weekday % 7]
		ws[f'B{row}'].value = hd.strftime('%d/%m/%Y')

		fqty = []
		mqty = []
		try:
			fqty = [int(x) for x in str(sh[f'N{srcrow + 6}'].value).split('/')]
		except:
			pass
		try:
			mqty = [int(x) for x in str(sh[f'N{srcrow + 7}'].value).split('/')]
		except:
			pass

		mx = max(len(fqty), len(mqty))
		fqty += [0] * (mx - len(fqty))
		mqty += [0] * (mx - len(mqty))
		
		ws[f'C{row}'].value = sum(fqty)
		ws[f'D{row}'].value = sum(mqty)

		ws[f'E{row}'].value = f'=ROUNDUP(((C{row}*1.05)+(D{row}*1.1))*2, 0)'
		ws[f'F{row}'].value = sh[f'N{srcrow + 4}'].value

		sexing_date = sh[f'N{srcrow + 3}'].value
		if not pd.isnull(sexing_date):
			if sexing_date.hour > 18:
				ws[f'G{row}'].value = sexing_date.strftime('%H:%m') + ' DO DIA ' + (hd - datetime.timedelta(days=1)).strftime('%d/%m/%Y')
			else:
				ws[f'G{row}'].value = sexing_date.strftime('%H:%m') + ' DO DIA ' + hd.strftime('%d/%m/%Y')


		ws[f'C{row}'].number_format = '#,##0'
		ws[f'D{row}'].number_format = '#,##0'
		ws[f'E{row}'].number_format = '#,##0'

		dispatch_date = sh[f'N{srcrow + 8}'].value
		if isinstance(dispatch_date, datetime.datetime):
			ws[f'H{row}'].value = dispatch_date.strftime('%d/%m/%Y')


		ws[f'I{row}'].value = sh[f'N{srcrow + 9}'].value

		custs = sh[f'N{srcrow + 5}'].value
		if pd.isnull(custs):
			custs = '-'
		else:
			custs = str(custs).upper()
			ws[f'J{row}'].value = custs.replace('\n', ' / ')
			set_range_style(ws, row, 1, row, 10, fill=color_fill("FFFFFF00"))
		row += 1


	
		females = []
		males = []

		for j in range(srcrow, srcrow + 45):
			try:
				if int(sh[f'L{j}'].value) > 0:
					if find_line(db, str(sh[f'B{j}'].value)) == 'F':
						females.append(sh[f'C{j}'].value)
					elif find_line(db, str(sh[f'B{j}'].value)) == 'M':
						males.append(sh[f'C{j}'].value)
			except: pass

		if custs != '-':
			siz = max(len(females), len(males))
			for j, x in enumerate(custs.replace('\n', ' / ').split('/')):

				strrow = (siz + 6) * j
				merge_range(lotes, strrow + 1, i * 3 + 1, strrow + 1, i * 3 + 2, hd.strftime('%d/%m/%Y'))
				merge_range(lotes, strrow + 2, i * 3 + 1, strrow + 2, i * 3 + 2, re.sub(r'\([^)]*\)', '', x))#custs))

				merge_range(lotes, strrow + 3, i * 3 + 1, strrow + 3, i * 3 + 2, 'LOTES')
				set_row(lotes, strrow + 4, [fqty[j], mqty[j]], i * 3 + 1)
				set_row(lotes, strrow + 5, ['FÊMEA', 'MACHO'], i * 3 + 1)

				set_range_style(lotes, strrow + 1, i * 3 + 1, strrow + 4, i * 3 + 2, font=Font(bold=True, color='FFFFFFFF'), fill=color_fill('FF000000'))
				set_cell_style(lotes.cell(strrow + 2, i * 3 + 1), alignment=Alignment(wrap_text = True))
				set_cell_style(lotes.cell(strrow + 5, i * 3 + 1), font=Font(bold=True, color='FFFFFFFF'), fill=color_fill('FFFF0000'))
				set_cell_style(lotes.cell(strrow + 5, i * 3 + 2), font=Font(bold=True, color='FFFFFFFF'), fill=color_fill('FF0070C0'))

				
				
				print(x)
				
				set_col(lotes, i * 3 + 1, females, strrow + 6)
				set_col(lotes, i * 3 + 2, males, strrow + 6)
				set_column_width(lotes, i * 3 + 1, 12)
				set_column_width(lotes, i * 3 + 2, 12)

	set_range_style(ws, 5, 1, 5, 10, font=Font(bold=True))
	set_range_style(ws, 8, 1, 9, 10, font=Font(bold=True))
	set_range_style(ws, 10, 1, 16, 1, font=Font(bold=True))
	set_range_style(ws, 17, 5, 17, 7, font=Font(bold=True))

	start_date = ws[f'B{row - 7}'].value
	end_date = ws[f'B{row - 1}'].value

	merge_range(ws, row - 10, 1, row - 10, 10, f"Semana: {start_date} a {end_date}")

	ws.title = f"{ws[f'B{row - 7}'].value} A {ws[f'B{row - 1}'].value}".replace('/', '-')

	ws[f'E{row}'].value = f'=ROUNDDOWN(SUM(E{row - 7}:E{row - 1}), 0)'
	ws[f'F{row}'].value = f'=ROUNDDOWN(E{row}/AVERAGE(F{row - 7}:F{row - 1}), 0)'
	ws[f'E{row}'].number_format = '#,##0'
	ws[f'F{row}'].number_format = '#,##0'
	ws[f'G{row}'].value = 'POR SEXADOR / SEMANA'

	set_range_border(ws, 8, 1, 9, 10)
	set_range_border(ws, 10, 1, 16, 10)
	set_range_border(ws, 17, 5, 17, 6)
	set_columns_width(ws, [16, 13, 11, 11, 11, 11, 23, 23, 23, 67])

	ws[f'E{row + 1}'].value = 'Aviagen América Latina'
	ws[f'E{row + 2}'].value = 'Incubatório de Matrizes'
	ws[f'E{row + 3}'].value = 'Rio Claro - SP'

	set_content_alignment(ws, h='center')

	set_content_alignment(lotes, h='center')
	set_content_border(lotes)


	autofit_all_columns(ws)

	lotes.sheet_view.zoomScale = 150

	#PRINT CONFIG

	"""img = drawing.image.Image('aviagen-logo.png')
	p2e = pixels_to_EMU
	h, w = img.height * (40 / img.height), img.width * (40 / img.height)
	position = XDRPoint2D(p2e(5), p2e(5))
	size = XDRPositiveSize2D(p2e(w), p2e(h))
	img.anchor = AbsoluteAnchor(pos=position, ext=size)
	#img.anchor = 'A1'
	ws.add_image(img)"""

	fit(ws)
	return wb
