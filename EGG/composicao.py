import pandas as pd
import math
import numpy as np
from pyx.xl import *
from pyx.matrix_utility import transpose
from pyx.dataframe_utility import pick, strall, row_of, fillnext, resum, find_rows, split, select, segment, isnull
from pyx.array_utility import concat
from datetime import datetime as dt
from openpyxl import Workbook, load_workbook

from EGG.hatchery import line, egg_height

# set the chained_assignment option
#pd.options.mode.chained_assignment = "raise" # raises exception in case of warning
#pd.options.mode.chained_assignment = "warn"  # prints warning in case of warning, no exception is raised
pd.options.mode.chained_assignment = None    # no warning message and no exception is raised


def str_strain_flock(df):
	df['MTECH_FLOCK_ID'] = df['MTECH_FLOCK_ID'].map(lambda x: '{:0>5}'.format(x))
	df['STRAIN_CODE'] = df['STRAIN_CODE'].map(lambda x: str(x))
	return df

def init_hatch_results(df, parse_dates=['PRODUCTION_DATE', 'HATCH_DATE']):
	for x in parse_dates:
		df[x] = pd.to_datetime(df[x], dayfirst=True)
	df = str_strain_flock(df)
	#df['EGG_HEIGHT'] = df['EGG_CLASS'].map(egg_height)
	if 'LINE' not in df.columns:
		df['LINE'] = df['STRAIN_CODE'].map(line)
	return df

def init_set_grid(df):
	df['HATCH DATE'] = pd.to_datetime(df['HATCH DATE'], dayfirst=True)
	df['STRAIN'] = df['STRAIN'].map(lambda x: str(x))
	df['DIVISION'] = df['DIVISION'].fillna('')
	df['DIVISION'] = df['DIVISION'].map(lambda x: str(x))
	df['LINE'] = df['STRAIN'].map(line)
	df = df[df['ORDER STATUS'] != 'Cancelled'].reset_index()
	df['PP'] = df['MALES'] + df['FEMALES']
	fillnext(df, ['ORDER NUMBER', 'CUSTOMER NAME', 'COUNTRY', 'HATCH DATE', 'CHXBOX'])
	return df


def fillgtas(sheet, gtas):
	if 'GTA_NUMBER' not in sheet.columns:
		sheet['GTA_NUMBER'] = ''
	gta = sheet.columns.get_loc('GTA_NUMBER')
	for i in range(sheet.shape[0]):
		#print(sheet.iloc[i, gta])
		if not isnull(sheet.iloc[i, gta]): continue
		print(sheet['MTECH_FLOCK_ID'][i])
		idx = row_of(gtas, { 'PRODUCTION_DATE': sheet['PRODUCTION_DATE'][i], 'MTECH_FLOCK_ID': sheet['MTECH_FLOCK_ID'][i], 'STRAIN_CODE': sheet['STRAIN_CODE'][i] })
		print(idx)
		if idx > -1: sheet.iloc[i, gta] = gtas['GTA_NUMBER'][idx]
		else: sheet.iloc[i, gta] = '-'
	return sheet






def flocks_from(df, line):
	df = df[df['LINE'] == line][['MTECH_FLOCK_ID', 'FARM_NAME']].drop_duplicates()
	return [ line ] + list(df.apply(lambda x: f"{x['MTECH_FLOCK_ID']} - {x['FARM_NAME']}", axis=1))


def flocks(ws, df):
	male_flocks = flocks_from(df, 'M')
	female_flocks = flocks_from(df, 'F')
	set_rng(ws, transpose([ female_flocks ] + [ male_flocks ]))
	set_content_border(ws)





def comp_ovos(ws, df, orders=None, vaccines=None): #ws is the worksheet where the hatch composition will be displayed coming from the dataframe df
	on =  df['ORDERNO'].iloc[0]
	custname = df[df['ORDERNO'] == on]['CUSTNAME'].unique()[0]
	
	df = df.assign(MTECH_FLOCK_ID=concat([x[4:] + "-" for x in df['FARM_CODE']], list(df['MTECH_FLOCK_ID'])))

	df = df[['FARM_NAME', 'GTA_NUMBER', 'LINE', 'STRAIN_CODE', 'MTECH_FLOCK_ID', 'EGGS' ]]#, 'FARM_CODE']]



	df = df.assign(EGGS=pd.to_numeric(df['EGGS']))

	df = df.groupby(by=['FARM_NAME', 'GTA_NUMBER', 'MTECH_FLOCK_ID']).sum(numeric_only = True)#, as_index=False).sum()#.reset_index() #it will be sorted by the given indices
	df.rename(columns = {'FARM_NAME': 'GRANJA', 'GTA_NUMBER': 'GTA', 'MTECH_FLOCK_ID': 'LOTE', 'EGGS': 'OVOS'}, inplace=True)

	df.rename_axis(['GRANJA', 'GTA', 'LOTE'], inplace=True)
	df = df.reset_index()
	#print(df.columns)
	

	merge_range(ws, 1, 4, 1, 4, custname, border=thick_border, font=Font(bold=True), fill=color_fill("FFFFC000"))
	merge_range(ws, 2, 4, 2, 4, f'ORDEM: {on}', border=thick_border, font=Font(bold=True))

	min_row = 1
	

	dfcpy(ws, df, min_row + 2, 1,	header_fill=color_fill("FF00B0F0"), header_font=Font(bold=True, color="FFFFFFFF"))

	last_row = min_row + 3 + (df.shape[0] - 1)

	set_range_style(ws, min_row + 3, 1, last_row, 2, font=Font(bold=True)) #BOLD FIRST AND SECOND COLUMN

	merge_vertically(ws, min_row + 2, 1, last_row, 2)
	set_range_border_by_group(ws, min_row + 2, 1, last_row, 4)
	
	set_row(ws, last_row + 1, total_row(min_row + 3, 4, last_row, 4), 4, fill=color_fill("FFFFFF00"), border=thick_border)


def comp_aves(ws, df, orders, vaccines): #ws is the worksheet where the hatch composition will be displayed coming from the dataframe df
	custname = df['CUSTNAME'].iloc[0]	#df[df['ORDERNO'] == on]['CUSTNAME'].unique()[0]
	on =  df['ORDERNO'].iloc[0]
	df = df.assign(MTECH_FLOCK_ID=concat([x[4:] + "-" for x in df['FARM_CODE']], list(df['MTECH_FLOCK_ID'])))

	df = df[['FARM_NAME', 'GTA_NUMBER', 'LINE', 'STRAIN_CODE', 'MTECH_FLOCK_ID', 'EGGS', 'EST_SALEABLE_QTY' ]]#, 'FARM_CODE']]

	df = df.assign(**{'PP MACHO': pd.to_numeric(df.loc[df['LINE'] == 'M']['EST_SALEABLE_QTY'])})
	df = df.assign(**{'PP FÊMEA': pd.to_numeric(df.loc[df['LINE'] == 'F']['EST_SALEABLE_QTY'])})
	df = df.assign(**{'PH MACHO': pd.to_numeric(np.ceil(df.loc[df['LINE'] == 'F']['EST_SALEABLE_QTY'] * 1.17))})
	df = df.assign(**{'PH FÊMEA': pd.to_numeric(np.ceil(df.loc[df['LINE'] == 'M']['EST_SALEABLE_QTY'] * 1.17))})
	df = df.drop(columns=['EST_SALEABLE_QTY'])	

	df = df.assign(EGGS=pd.to_numeric(df['EGGS']))

	df = df.groupby(by=['FARM_NAME', 'GTA_NUMBER', 'MTECH_FLOCK_ID']).sum(numeric_only = True)#, as_index=False).sum()#.reset_index() #it will be sorted by the given indices
	df.rename(columns = {'FARM_NAME': 'GRANJA', 'GTA_NUMBER': 'GTA', 'MTECH_FLOCK_ID': 'LOTE', 'EGGS': 'OVOS'}, inplace=True)
	df.rename(columns = {'PP FÊMEA': 'FÊMEA', 'PP MACHO': 'MACHO', 'PH FÊMEA': 'FÊMEA', 'PH MACHO': 'MACHO'}, inplace=True)
	df.rename_axis(['GRANJA', 'GTA', 'LOTE'], inplace=True)
	df = df.reset_index()
	#print(df.columns)
	

	merge_range(ws, 1, 5, 1, 6, custname, border=thick_border)
	idx = row_of(orders, { 'ORDER NUMBER': on })
	min_row = 1
	if idx > -1:
		vaccine_numbers = doencas(orders.iloc[idx, orders.columns.get_loc('VACCINES')], vaccines)
		set_col(ws, 5, vaccine_numbers, 2, ignore_merge=False, scale=(1, 2), border=thin_border)
		min_row = len(vaccine_numbers) + 1

	
	min_row -= 1 #ISSO É POG
	#merge_range(ws, min_row + 1, 5, min_row + 1, 6, custname, border=thick_border)
	merge_range(ws, min_row + 2, 5, min_row + 2, 8, "PINTOS", fill=color_fill("FFFFFF00"), border=thick_border)
	set_row(ws, min_row + 3, ['MATRIZ', 'PH'], 5, ignore_merge=False, scale=(1, 2), fill=color_fill("FFFFFF00"), border=thick_border)
	dfcpy(ws, df, min_row + 4, 1	, header_fill=color_fill("FF00B0F0"), header_font=Font(bold=True, color="FFFFFFFF"))

	last_row = min_row + 5 + (df.shape[0] - 1)

	set_range_style(ws, min_row + 5, 1, last_row, 2, font=Font(bold=True)) #BOLD FIRST AND SECOND COLUMN

	merge_vertically(ws, min_row + 5, 1, last_row, 2)
	set_range_border_by_group(ws, min_row + 5, 1, last_row, 8)
	
	set_row(ws, last_row + 1, total_row(min_row + 5, 4, last_row, 8), 4	, fill=color_fill("FFFFFF00"), border=thick_border)

	merge_range(ws, last_row + 2, 5, last_row + 2, 6, "=" + ws.cell(last_row + 1, 5).coordinate + "+" + ws.cell(last_row + 1, 6).coordinate, fill=color_fill('FFFFC000'), border=thick_border)
	merge_range(ws, last_row + 2, 7, last_row + 2, 8, "=" + ws.cell(last_row + 1, 7).coordinate + "+" + ws.cell(last_row + 1, 8).coordinate, fill=color_fill('FFFFC000'), border=thick_border)
	merge_range(ws, last_row + 1, 3, last_row + 1, 3, "TOTAL", border=thick_border)
	set_range_border(ws, min_row + 4, 1, min_row + 4, 8)

def doencas(vaccine_numbers, vaccines):
	result = []
	vaccine_numbers = str(vaccine_numbers)
	if not isnull(vaccine_numbers):
		#print(vaccine_numbers)
		result = [vaccines.iloc[int(x), vaccines.columns.get_loc('Doenças')] for x in vaccine_numbers.replace(' ', '').replace('.', ',').split(',')]
		result = list(set(result))
	return result

def divideit(df, orders):
	hd = pd.to_datetime(df['HATCH_DATE'].iloc[0]).date()
	on = df['ORDERNO'].iloc[0]
	cn = df['CUSTNAME'].iloc[0]

	if orders[orders['ORDER NUMBER'] == on].shape[0] > 0:
		cn = orders[orders['ORDER NUMBER'] == on].iloc[0, orders.columns.get_loc('CUSTOMER NAME')]
		#cns = orders[(orders['HATCH DATE'].strftime("%d-%m-%Y") == hd.strftime("%d-%m-%Y")) & (orders['CUSTOMER NAME'] == cn)]
		cns = orders[orders['CUSTOMER NAME'] == cn]
		cns['HATCH DATE'] = pd.to_datetime(cns['HATCH DATE']).dt.date#, errors='coerce')
			
		cns = cns[cns['HATCH DATE'] == hd]
		#print(cns)
		

		if cns.shape[0] > 0:
			orders1 = []
			for st in cns['LINE'].unique():
				orders2 = [ df[df['LINE'] == st] ]
				cns2 = cns[(cns['LINE'] == st) & ((cns['ORDER NUMBER'] == on) | (cns['EGGS'] == 0))]		#17/08/2023

				total_ordered = cns2['PP'].sum()
				total = orders2[0]['EST_SALEABLE_QTY'].sum()

				for i in range(1, cns2.shape[0]):
					if cns2.iloc[i, cns2.columns.get_loc('EST SAL CHICKS')] != 0: continue	#FAÇA UMA FUNÇÃO PARA SIMPLIFICAR
						
						
					term = math.floor((cns2.iloc[i, cns2.columns.get_loc('PP')] / total_ordered) * total)

					picked = pick(orders2[0], term, 'EST_SALEABLE_QTY', ['EGGS'])
					orders2[0] = picked[1]

					orders2.append(picked[0])
					orders2[i] = orders2[i].assign(CUSTNAME=[x + " " + str(i + 1) for x in orders2[i]['CUSTNAME']]) #USE ORDER NUMBER LATER
					
				
				#APENAS PARA EXPORTAÇÕES
				for i in range(0, cns2.shape[0]):
					
					if cns2.iloc[i, cns2.columns.get_loc('COUNTRY')] != 'BR':
						print(orders2[i][['GTA_NUMBER', 'EST_SALEABLE_QTY']])
						orders2[i] = resum(orders2[i], cns2.iloc[i, cns2.columns.get_loc('PP')], 'EST_SALEABLE_QTY')
						print(orders2[i][['GTA_NUMBER', 'EST_SALEABLE_QTY']])
				orders1 += orders2
				#orders1.append(pd.concat(orders2))

			#return orders1
			df = pd.concat(orders1, ignore_index=True)


	if orders[orders['ORDER NUMBER'] == on].shape[0] > 0:
		df3 = orders[orders['ORDER NUMBER'] == on]
		divs = []
		for ln in df3['LINE'].unique():
			df4 = df3[df3['LINE'] == ln]
			div = df4.iloc[0, orders.columns.get_loc('DIVISION')]
			#print(div)
			if len(div) > 0 and div[0] == '/': #DIVISÃO ADIANTADA PELA PLANILHA DE CONTROLE DE ENTREGAS
				
				dfs = split(df[df['LINE'] == ln], int(float(div[1:])), 'EST_SALEABLE_QTY', ['EGGS'])
			else: #DIVISÃO POR QUANTIDADES ESPECIFICAS OU SEM DIVISÃO
				div = div.replace('  ', ' ').split(' ')
				div = list([int(float(x)) for x in div if x != ''])
				#print(div)
				dfs = split(df[df['LINE'] == ln], div, 'EST_SALEABLE_QTY', ['EGGS'])
			for i in range(len(dfs)):
				dfs[i]['CUSTNAME'] = dfs[i]['CUSTNAME'].map(lambda x: x + '' + str(i + 1))

			divs += dfs
			#divs.append(pd.concat(dfs, ignore_index=True))
		return divs

	#NO DIVISION
	return [df]








def inserir_vacinas_ph(ws, orders, date, vaccines, min_col = 4):
	by_products = orders[(orders['HATCH DATE'].dt.date == date) & (orders['ORDER TYPE'] == 'B')][['CUSTOMER NAME', 'VACCINES']]
	#print(by_products)
	for i in range(by_products.shape[0]):
		col = min_col + (i * 2)
		custname = by_products.iloc[i, by_products.columns.get_loc('CUSTOMER NAME')]
		if custname == 'COOPERATIVA PECUÁRIA':
			custname = 'HOLAMBRA'
		merge_range(ws, 1, col, 1, col, custname, fill=color_fill("FFFFFF00"), border=thick_border)

		vaccine_numbers = doencas(by_products.iloc[i, by_products.columns.get_loc('VACCINES')], vaccines)
		set_col(ws, col, vaccine_numbers, 2, ignore_merge=False, border=thin_border)

	return ws

def gerar_composicao(hatch, orders, vaccines, gtas, comp2xl):
	hatch = init_hatch_results(hatch)
	orders = init_set_grid(orders)
	gtas = str_strain_flock(gtas)
	#hatch = hatch.sort_values(by=['HATCH_DATE', 'ORDERNO', 'CUSTNAME', 'EGG_HEIGHT', 'FLOCK_AGE', 'GTA_NUMBER', 'FARM_CODE', 'MTECH_FLOCK_ID', 'STRAIN_CODE'])
	hatch = hatch.sort_values(by=['HATCH_DATE', 'ORDERNO', 'CUSTNAME', 'FARM_NAME', 'GTA_NUMBER', 'FARM_CODE', 'MTECH_FLOCK_ID', 'STRAIN_CODE'])

	
	gtas['PRODUCTION_DATE'] = gtas['PRODUCTION_DATE'].map(lambda x: pd.to_datetime(x).strftime('%Y-%m-%d'))
	hatch['PRODUCTION_DATE'] = hatch['PRODUCTION_DATE'].map(lambda x: pd.to_datetime(x).strftime('%Y-%m-%d'))

	hatch = fillgtas(hatch, gtas)
	
	result = []
	for df in segment(hatch, ['HATCH_DATE', 'ORDERNO']):
		result += divideit(df, orders)
	hatch = pd.concat(result, ignore_index=True)
	
	wb = EmptyWorkbook()

	date_column = 'HATCH_DATE' if comp2xl.__name__ == 'comp_aves' else 'SHIP_DATE'
	title = 'COMPOSIÇÃO DE ' + ('NASCIMENTO' if comp2xl.__name__ == 'comp_aves' else 'EXPORTAÇÃO DE OVOS')

	for hd in hatch[date_column].unique():
		strhd = pd.to_datetime(hd, dayfirst=True).date()
	
		ws = wb.create_sheet(title=strhd.strftime("%d-%m-%Y")) #str(hd)[:10])
	
		matrices = [[], [], [], []]
		matrices[0].append(Worksheet())
		merge_range(matrices[0][0], 1, 1, 1, 8, title, fill=color_fill("FFFFFF00"), border=thick_border)
		matrices[1].append(Worksheet())
		set_row(matrices[1][0], 1, ["DATA", strhd.strftime("%d/%m/%Y")])#str(hd)[:10]])

		if comp2xl.__name__ == 'comp_aves':
			inserir_vacinas_ph(matrices[1][0], orders, strhd, vaccines)


		set_range_border(matrices[1][0], 1, 1, 1, 2)
		order_count = 0
		for on in hatch[hatch[date_column] == hd]['ORDERNO'].unique():
			report = hatch[hatch['ORDERNO'] == on]
		
			#print(on)
			for cn in report['CUSTNAME'].unique():
				df3 = report[report['CUSTNAME'] == cn]
				matrices[2].append(Worksheet())
				matrices[3].append(Worksheet())
				flocks(matrices[2][0 + order_count], df3)
				comp2xl(matrices[3][0 + order_count], df3, orders, vaccines)
				order_count += 1


		paste_worksheet_matrix(ws, matrices)
		set_content_alignment(ws, v='top')
		autofit_all_columns(ws)
		set_columns_width(ws, 23)

	print("Done!")
	return [wb, hatch]
	
